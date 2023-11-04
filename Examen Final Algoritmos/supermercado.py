import openpyxl

# Crear un nuevo libro de Excel
workbook = openpyxl.Workbook()

sheet = workbook.active
sheet.title = "Lista de Supermercados" 

sheet['A1'] = "Supermercado"  
sheet['B1'] = "Producto" 
sheet['C1'] = "Precio"  

supermarkets = [
    ("Supermercado A", "Manzanas", 2.99),
    ("Supermercado A", "Peras", 3.49),
    ("Supermercado B", "Naranjas", 2.79),
    ("Supermercado B", "Kiwi", 3.29),
]

for row, data in enumerate(supermarkets, start=2):
    sheet.cell(row=row, column=1, value=data[0])
    sheet.cell(row=row, column=2, value=data[1])
    sheet.cell(row=row, column=3, value=data[2])

workbook.save("lista_supermercados.xlsx")

